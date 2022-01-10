# -*- coding: utf-8 -*-
# ProjectName:  ZhiHu-Spider
# FileName:     save_data
# Description:  
# CreateDate:   2021/12/25

import os

from openpyxl import Workbook, load_workbook

from database import MysqlClient
from database import MongoDBClient
from handler import ConfigHandler


class SaveData(object):

    def __init__(self, data, table_name, unique_id):
        self.data = data
        self.table_name = table_name
        self.unique_id = unique_id
        self.conf = ConfigHandler()

    def to_mysql(self):
        for item in self.data:
            mysql_client = MysqlClient()
            mysql_client.save(item, table_name=self.table_name)

    def to_mongo(self):
        for item in self.data:
            mongo_client = MongoDBClient()
            db = mongo_client.db
            db[self.table_name].update_one({self.unique_id: item[self.unique_id]}, {'$set': item}, upsert=True)

    def to_xlsx(self):
        for item in self.data:
            file_name = 'data/{}.xlsx'.format(self.table_name, self.table_name)
            try:
                wb = load_workbook(file_name)
                ws = wb['Sheet']
            except:
                wb = Workbook()
                ws = wb['Sheet']
                my_keys = list(item.keys())
                ws.append(my_keys)
            my_values = list(item.values())
            ws.append(my_values)
            wb.save(file_name)

    def run(self):
        if self.conf.use_mysql == 1:
            self.to_mysql()
        if self.conf.use_mongo == 1:
            self.to_mongo()
        if self.conf.use_xlsx == 1:
            if not os.path.exists('data'.format(self.table_name)):
                try:
                    os.makedirs('data'.format(self.table_name))
                except FileExistsError:
                    pass
            self.to_xlsx()
